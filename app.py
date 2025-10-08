"""
Face Recognition Attendance System - Web Version
Flask Backend with REST API
"""

from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
import cv2
import numpy as np
# ... NO face_recognition import!
from deepface import DeepFace
import csv
import os
from datetime import datetime, date
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pickle
import base64
import io
from PIL import Image

app = Flask(__name__)
CORS(app)

class AttendanceSystem:
    def __init__(self):
        self.base_dir = Path.cwd()
        self.photos_dir = self.base_dir / "photos"
        self.attendance_dir = self.base_dir / "attendance"
        self.registration_file = self.base_dir / "registration.csv"
        self.encodings_file = self.base_dir / "face_encodings.pkl"
        self.attendance_cache_file = self.base_dir / "attendance_cache.pkl"
        
        # Create directories
        self.photos_dir.mkdir(exist_ok=True)
        self.attendance_dir.mkdir(exist_ok=True)
        
        # Initialize data
        self.known_face_encodings = []
        self.known_face_data = []
        self.today_attended = self.load_attendance_cache()
        
        # Initialize CSV
        if not self.registration_file.exists():
            with open(self.registration_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Employee_ID', 'Name', 'Phone', 'Address', 'Registration_Date'])
        
        self.load_face_encodings()
        
        # Camera state
        self.camera = None
    
    def load_attendance_cache(self):
        if self.attendance_cache_file.exists():
            with open(self.attendance_cache_file, 'rb') as f:
                cache = pickle.load(f)
                if cache.get('date') == str(date.today()):
                    return cache.get('attended', set())
        return set()
    
    def save_attendance_cache(self):
        cache = {'date': str(date.today()), 'attended': self.today_attended}
        with open(self.attendance_cache_file, 'wb') as f:
            pickle.dump(cache, f)
    
    def load_face_encodings(self):
        if self.encodings_file.exists():
            with open(self.encodings_file, 'rb') as f:
                data = pickle.load(f)
                self.known_face_encodings = data['encodings']
                self.known_face_data = data['data']
    
    def save_face_encodings(self):
        data = {'encodings': self.known_face_encodings, 'data': self.known_face_data}
        with open(self.encodings_file, 'wb') as f:
            pickle.dump(data, f)
    
    def register_employee(self, emp_id, name, phone, address, image_data):
        """Register employee with base64 image"""
        # Check if exists
        if self.registration_file.exists():
            with open(self.registration_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row['Employee_ID'] == emp_id:
                        return {'success': False, 'message': f'Employee ID {emp_id} already exists'}
        
        # Decode image
        try:
            image_bytes = base64.b64decode(image_data.split(',')[1])
            image = Image.open(io.BytesIO(image_bytes))
            image_np = np.array(image)
            
            # Convert RGB to BGR for OpenCV
            if len(image_np.shape) == 3 and image_np.shape[2] == 3:
                image_bgr = cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR)
            else:
                image_bgr = image_np
        except Exception as e:
            return {'success': False, 'message': f'Invalid image data: {str(e)}'}
        
        # Detect face
        rgb_image = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_image)
        
        if len(face_locations) == 0:
            return {'success': False, 'message': 'No face detected in image'}
        
        # Generate encoding
        face_encodings = face_recognition.face_encodings(rgb_image, face_locations)
        if len(face_encodings) == 0:
            return {'success': False, 'message': 'Could not generate face encoding'}
        
        face_encoding = face_encodings[0]
        
        # Save photo
        photo_filename = f"{emp_id}_{name.replace(' ', '_')}.jpg"
        photo_path = self.photos_dir / photo_filename
        cv2.imwrite(str(photo_path), image_bgr)
        
        # Add to known faces
        self.known_face_encodings.append(face_encoding)
        self.known_face_data.append({
            'emp_id': emp_id,
            'name': name,
            'phone': phone,
            'address': address
        })
        self.save_face_encodings()
        
        # Save to CSV
        with open(self.registration_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([emp_id, name, phone, address, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        return {'success': True, 'message': 'Registration successful', 'emp_id': emp_id, 'name': name}
    
    def mark_attendance(self, emp_data, timestamp):
        today = date.today()
        year_month = today.strftime('%Y-%m')
        month_dir = self.attendance_dir / year_month
        month_dir.mkdir(exist_ok=True)
        
        excel_file = month_dir / f"attendance_{today.strftime('%Y-%m-%d')}.xlsx"
        
        if excel_file.exists():
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        
        sheet_name = f"{emp_data['emp_id']}_{emp_data['name'][:20]}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            headers = ['Date', 'Time', 'Employee ID', 'Name', 'Status']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        ws.append([
            today.strftime('%Y-%m-%d'),
            timestamp.strftime('%H:%M:%S'),
            emp_data['emp_id'],
            emp_data['name'],
            'Present'
        ])
        
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(excel_file)
        return str(excel_file)
    
    def mark_unknown_attendance(self, timestamp):
        today = date.today()
        year_month = today.strftime('%Y-%m')
        month_dir = self.attendance_dir / year_month
        month_dir.mkdir(exist_ok=True)
        
        excel_file = month_dir / f"attendance_{today.strftime('%Y-%m-%d')}.xlsx"
        
        if excel_file.exists():
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        
        sheet_name = "Unknown_Faces"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            headers = ['Date', 'Time', 'Status']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        ws.append([today.strftime('%Y-%m-%d'), timestamp.strftime('%H:%M:%S'), 'Unknown Face Detected'])
        wb.save(excel_file)
    
    def process_frame(self, frame):
        """Process a single frame for face recognition"""
        # Resize for faster processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        
        results = []
        
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            
            name = "Unknown"
            emp_id = None
            attended = False
            
            if len(self.known_face_encodings) > 0:
                matches = face_recognition.compare_faces(self.known_face_encodings, face_encoding, tolerance=0.6)
                face_distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)
                
                if len(face_distances) > 0:
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        emp_data = self.known_face_data[best_match_index]
                        name = emp_data['name']
                        emp_id = emp_data['emp_id']
                        
                        if emp_id not in self.today_attended:
                            timestamp = datetime.now()
                            self.mark_attendance(emp_data, timestamp)
                            self.today_attended.add(emp_id)
                            self.save_attendance_cache()
                            attended = True
            
            if name == "Unknown":
                self.mark_unknown_attendance(datetime.now())
            
            results.append({
                'name': name,
                'emp_id': emp_id,
                'box': {'top': top, 'right': right, 'bottom': bottom, 'left': left},
                'attended': attended,
                'already_attended': emp_id in self.today_attended if emp_id else False
            })
        
        return results
    
    def get_all_employees(self):
        """Get list of all registered employees"""
        employees = []
        if self.registration_file.exists():
            with open(self.registration_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    employees.append(row)
        return employees

# Initialize system
system = AttendanceSystem()

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    result = system.register_employee(
        data['emp_id'],
        data['name'],
        data['phone'],
        data['address'],
        data['image']
    )
    return jsonify(result)

@app.route('/api/employees', methods=['GET'])
def get_employees():
    employees = system.get_all_employees()
    return jsonify({'employees': employees, 'count': len(employees)})

@app.route('/api/process-frame', methods=['POST'])
def process_frame():
    """Process a frame from webcam"""
    data = request.json
    try:
        image_data = data['frame']
        image_bytes = base64.b64decode(image_data.split(',')[1])
        image = Image.open(io.BytesIO(image_bytes))
        frame = np.array(image)
        
        # Convert RGB to BGR
        if len(frame.shape) == 3 and frame.shape[2] == 3:
            frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
        
        results = system.process_frame(frame)
        return jsonify({'success': True, 'faces': results})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/attendance-today', methods=['GET'])
def attendance_today():
    """Get today's attendance list"""
    attended_ids = list(system.today_attended)
    attended_employees = []
    
    for emp_data in system.known_face_data:
        if emp_data['emp_id'] in attended_ids:
            attended_employees.append(emp_data)
    
    return jsonify({
        'count': len(attended_employees),
        'employees': attended_employees,
        'date': str(date.today())
    })

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get system statistics"""
    total_employees = len(system.known_face_data)
    today_attendance = len(system.today_attended)
    attendance_rate = (today_attendance / total_employees * 100) if total_employees > 0 else 0
    
    return jsonify({
        'total_employees': total_employees,
        'today_attendance': today_attendance,
        'attendance_rate': round(attendance_rate, 1),
        'date': str(date.today())
    })

if __name__ == '__main__':
    # Create templates directory
    templates_dir = Path('templates')
    templates_dir.mkdir(exist_ok=True)
    
    print("="*50)
    print("🌐 Face Recognition Attendance Web System")
    print("="*50)
    print(f"📂 Photos: {system.photos_dir}")
    print(f"📊 Attendance: {system.attendance_dir}")
    print(f"👥 Registered Employees: {len(system.known_face_data)}")
    print("="*50)
    print("🚀 Starting server at http://localhost:5000")
    print("="*50)
    
    app.run(host='0.0.0.0', port=5000, debug=True, threaded=True)
